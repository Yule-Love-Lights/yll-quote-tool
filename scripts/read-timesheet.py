"""
Read Jason's Time Tracker spreadsheet into JSON — ledger row 507, step one.

    python scripts/read-timesheet.py "C:\\path\\to\\Time Tracker.xlsx" > timesheet.json

WHY PYTHON, AND WHY A SEPARATE STEP. The meaning of this file lives in CELL
COLOUR — green on the Hours cell means those hours were already paid, no fill
means still owed, yellow on the Date cell marks the first day of a new pay
rate. Nothing else in the sheet records any of it, so a CSV export would
silently turn a year of settled history into a year of outstanding wages.

Reading colour out of an .xlsx needs a real xlsx library. openpyxl is already
available here; the node side would need a new dependency in package.json for
one throwaway script, on a SHARED file. So the parsing lives here and the
importing lives in scripts/import-timesheet.ts, which drives the app's own
functions against the JSON this emits.

The JSON is also the point at which a human can check what was read before
anything is written: it is a plain, diffable record of how every row of the
spreadsheet was interpreted.
"""

import json
import sys
from datetime import datetime, timedelta

import openpyxl

SHEET = "Time Log"
PAID_FILL = "FF92D050"      # green  -> these hours were already paid
BOUNDARY_FILL = "FFFFFF00"  # yellow -> first day of a new pay rate
FIRST_DATA_ROW = 4          # rows 1-3 are the title, the tip and the headers

# Cells whose fill is a theme or indexed colour rather than plain rgb. Read
# argb() for why these are collected instead of shrugged at.
NON_RGB_FILLS = []


def argb(cell):
    """The cell's solid fill colour as AARRGGBB, or None if it has no fill.

    A fill that is NOT plain rgb -- a theme colour or an indexed one -- comes
    back as None, which is the same answer as "no fill at all". That silence is
    the dangerous case: a paid day whose green came from the theme palette
    would read as unpaid and put settled work back on the books. So non-rgb
    fills are collected in NON_RGB_FILLS and the caller refuses rather than
    guessing. This file's actual green (FF92D050) and yellow (FFFFFF00) are
    both plain rgb, so nothing is refused today -- but "today" is doing a lot
    of work in that sentence if the sheet is ever re-styled.
    """
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    if fg.type == "theme":
        # Theme 0 at tint 0 is "Background 1" — the sheet's own white. It is
        # not a colour signal, and this file really does carry it on three
        # date cells (A144-A146, the 2-4 Sep rows), which are correctly NOT
        # rate boundaries. Everything else themed is genuinely ambiguous.
        if fg.theme == 0 and not fg.tint:
            return None
        NON_RGB_FILLS.append((cell.coordinate, f"theme {fg.theme} tint {fg.tint}"))
        return None
    if fg.type != "rgb" or not fg.rgb:
        NON_RGB_FILLS.append((cell.coordinate, fg.type))
        return None
    return fg.rgb


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: read-timesheet.py <workbook.xlsx>", file=sys.stderr)
        return 2
    path = sys.argv[1]

    # Two loads on purpose: values gives the computed Hours, the default load
    # keeps the styling that carries the paid/unpaid meaning.
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_style = openpyxl.load_workbook(path)
    if SHEET not in wb_values.sheetnames:
        print(f"workbook has no sheet named {SHEET!r}", file=sys.stderr)
        return 1
    ws_v, ws_s = wb_values[SHEET], wb_style[SHEET]

    rows = []
    for r in range(FIRST_DATA_ROW, ws_v.max_row + 1):
        date_cell = ws_v.cell(r, 1)
        if not isinstance(date_cell.value, datetime):
            continue

        duration = ws_v.cell(r, 2).value
        seconds = int(round(duration.total_seconds())) if isinstance(duration, timedelta) else 0

        rows.append(
            {
                "sheetRow": r,
                "day": date_cell.value.date().isoformat(),
                "durationSeconds": seconds,
                "paid": argb(ws_s.cell(r, 3)) == PAID_FILL,
                "rateBoundary": argb(ws_s.cell(r, 1)) == BOUNDARY_FILL,
            }
        )

    # A sheet that yields no rows is a changed file or a wrong path, not an
    # empty history — fail rather than emit an empty import.
    if not rows:
        print(f"no dated rows found in {SHEET!r} from row {FIRST_DATA_ROW}", file=sys.stderr)
        return 1

    # REFUSE rather than guess. A theme-coloured fill on a Date or Hours cell
    # is indistinguishable here from no fill, and reading it as "no fill" means
    # reading a paid day as unpaid — the one mistake that puts settled work
    # back on the books.
    if NON_RGB_FILLS:
        where = ", ".join(f"{c} ({t})" for c, t in NON_RGB_FILLS[:8])
        print(
            f"{len(NON_RGB_FILLS)} cell(s) have a theme or indexed fill rather than a plain "
            f"colour, so whether they are green cannot be told from here: {where}. "
            "Re-fill them with a standard colour, or teach argb() to resolve the theme palette.",
            file=sys.stderr,
        )
        return 1

    json.dump(
        {
            "source": path,
            "sheet": SHEET,
            "readAt": datetime.now().astimezone().isoformat(),
            "rows": rows,
        },
        sys.stdout,
        indent=2,
    )
    print(file=sys.stdout)
    print(
        f"read {len(rows)} dated rows: "
        f"{sum(1 for x in rows if x['paid'])} paid, "
        f"{sum(1 for x in rows if not x['paid'])} unpaid, "
        f"{sum(1 for x in rows if x['rateBoundary'])} rate boundaries",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
